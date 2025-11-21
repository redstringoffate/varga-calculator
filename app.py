import sys, os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE_DIR)
sys.path.append(os.path.join(BASE_DIR, "utils"))

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.nakshatra_purushartha import NAKSHATRA_PURUSHARTHA


# ----------------------------------------------------------
# Utils
# ----------------------------------------------------------
from utils.sign import SIGNS, get_sign_symbol
from utils.conversion import to_decimal, to_degree_minute
from utils.varga_master import compute_vargas
from utils.nakshatra import get_nakshatra
from utils.nakshatra_lord import NAKSHATRA_LORD
from utils.purushartha import get_purushartha
from utils.color_map import (
    ELEMENT_COLOR,
    GRAHA_COLOR,
    PURUSHARTHA_COLOR,
    SIGN_TO_ELEMENT
)

# Glyph -> Sign name
GLYPH_TO_SIGN = {
    "♈":"Aries","♉":"Taurus","♊":"Gemini","♋":"Cancer",
    "♌":"Leo","♍":"Virgo","♎":"Libra","♏":"Scorpio",
    "♐":"Sagittarius","♑":"Capricorn","♒":"Aquarius","♓":"Pisces"
}

# ----------------------------------------------------------
# UI Setup
# ----------------------------------------------------------
st.set_page_config(page_title="Varga Chart Calculator", layout="wide")
st.title("🪷 Parashara Varga Calculator – Varga + Nakshatra + Purushartha + Colors")

PLANET_LIST = [
    "Sun","Moon","Mercury","Venus","Mars",
    "Jupiter","Saturn","Ketu","Rahu"
]

# ----------------------------------------
# Session Init
# ----------------------------------------
if "points" not in st.session_state:
    st.session_state.points = {}

for p in PLANET_LIST:
    if p not in st.session_state.points:
        st.session_state.points[p] = None

# ----------------------------------------
# ASC 입력
# ----------------------------------------
st.subheader("🔰 Ascendant 입력")
asc_c1, asc_c2, asc_c3 = st.columns([1.3, 1, 1])

with asc_c1:
    asc_sign = st.selectbox("Asc Sign", SIGNS, key="asc_sign")

with asc_c2:
    asc_deg = st.number_input("Degree", 0, 29, value=0, key="asc_deg")

with asc_c3:
    asc_min = st.number_input("Minute", 0, 59, value=0, key="asc_min")

# ----------------------------------------
# 행성 입력
# ----------------------------------------
st.subheader("📍 행성 위치 입력")

pc1, pc2, pc3, pc4 = st.columns([1.2, 1, 1, 1])

with pc1:
    selected_planet = st.selectbox("행성", PLANET_LIST, key="planet_sel")

with pc2:
    planet_sign = st.selectbox("Sign", SIGNS, key="planet_sign")

with pc3:
    planet_deg = st.number_input("Degree ", 0, 29, value=0, key="planet_deg")

with pc4:
    planet_min = st.number_input("Minute ", 0, 59, value=0, key="planet_min")

if st.button("💾 행성 위치 저장"):
    dec = to_decimal(planet_deg, planet_min)
    st.session_state.points[selected_planet] = (planet_sign, dec)
    st.success(f"{selected_planet} 저장 완료!")

# ----------------------------------------
# 현재 입력된 행성 표시
# ----------------------------------------
st.subheader("📌 현재 입력된 행성 위치")

for pl in PLANET_LIST:
    if st.session_state.points[pl] is None:
        st.write(f"- **{pl}**: ❌ 아직 입력 안됨")
    else:
        sign, dec = st.session_state.points[pl]
        d, m = to_degree_minute(dec)
        st.write(f"- **{pl}**: {get_sign_symbol(sign)} {sign} {d}°{m:02d}'")

# --------------------------------------------------------
# Varga 계산
# --------------------------------------------------------
st.markdown("---")

if st.button("🔮 Varga 계산하기"):

    missing = [p for p in PLANET_LIST if st.session_state.points[p] is None]
    if missing:
        st.error(f"❗ 입력되지 않은 행성: {', '.join(missing)}")
        st.stop()

    asc_decimal = to_decimal(asc_deg, asc_min)
    planets = {"ASC": (asc_sign, asc_decimal)}
    planets.update(st.session_state.points)

    varga_data = compute_vargas(planets)

    st.success("Varga 계산 완료!")

    labels = ["ASC"] + PLANET_LIST
    rows = []

    # -------------------------------
    # 테이블 구성 (Main / NK / PU)
    # -------------------------------
    for dn, pdata in varga_data.items():

        # 1) Main row
        row_main = {"Chart": dn}
        for lbl in labels:
            sign = pdata[lbl]["sign"]
            deg  = pdata[lbl]["degree"]
            house = pdata[lbl]["house"]

            d, m = to_degree_minute(deg)
            sym = get_sign_symbol(sign)
            row_main[lbl] = f"{sym} {d}°{m:02d}' ({house}H)"

        rows.append(row_main)

        # 2) NK row
        row_nk = {"Chart": f"{dn}_NK"}
        for lbl in labels:
            sign = pdata[lbl]["sign"]
            deg  = pdata[lbl]["degree"]
            nk_name, nk_pada = get_nakshatra(sign, deg)
            row_nk[lbl] = f"{nk_name}-{nk_pada}"
        rows.append(row_nk)

        # 3) PU row
        row_pu = {"Chart": f"{dn}_PU"}
        for lbl in labels:
            sign = pdata[lbl]["sign"]
            deg  = pdata[lbl]["degree"]
            nk_name, nk_pada = get_nakshatra(sign, deg)

            row_pu[lbl] = get_purushartha(nk_name, nk_pada)
        rows.append(row_pu)

    df = pd.DataFrame(rows)
    st.subheader("📊 Varga + Nakshatra + Purushartha")
    st.dataframe(df, use_container_width=True)

    # --------------------------------------------------------
    # Excel 생성 (openpyxl)
    # --------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Varga"

    # DF → Sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Column width
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    # -------------------------------
    # 색상 적용
    # -------------------------------
    for row in ws.iter_rows(min_row=2):
        chart_label = row[0].value

        # -----------------------
        # Main Row (element color)
        # -----------------------
        if "_NK" not in chart_label and "_PU" not in chart_label:
            for cell in row[1:]:
                v = cell.value
                if not v: 
                    continue

                glyph = v.split()[0]  # e.g., ♉
                sign = GLYPH_TO_SIGN.get(glyph)
                if not sign:
                    continue

                element = SIGN_TO_ELEMENT[sign]
                hexcode = ELEMENT_COLOR[element]

                cell.fill = PatternFill(
                    start_color=hexcode,
                    end_color=hexcode,
                    fill_type="solid"
                )

        # -----------------------
        # NK Row (lord color)
        # -----------------------
        if "_NK" in chart_label:
            for cell in row[1:]:
                v = cell.value
                if not v or "-" not in v:
                    continue
                nk, pada = v.split("-")

                lord = NAKSHATRA_LORD[nk]
                hexcode = GRAHA_COLOR.get(lord)

                if hexcode:
                    cell.fill = PatternFill(
                        start_color=hexcode,
                        end_color=hexcode,
                        fill_type="solid"
                    )

        # -----------------------
        # PU Row (font color)
        # -----------------------
        if "_PU" in chart_label:
            for cell in row[1:]:
                puru_full = cell.value
                puru = puru_full.split("-")[0]  # Dharma-K → Dharma

                hexcode = PURUSHARTHA_COLOR.get(puru)
                if hexcode:
                    cell.font = Font(color=hexcode, bold=True)

    # -------------------------------
    # Glossary sheet
    # -------------------------------
    ws2 = wb.create_sheet("Glossary")
    ws2.append(["Category", "Name", "Hex"])

    for elem, hexcode in ELEMENT_COLOR.items():
        ws2.append(["Element", elem, hexcode])

    for lord, hexcode in GRAHA_COLOR.items():
        ws2.append(["Nakshatra Lord", lord, hexcode or "None"])

    for p, hexcode in PURUSHARTHA_COLOR.items():
        ws2.append(["Purushartha", p, hexcode])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 15

    # -------------------------------
    # Save & Download
    # -------------------------------
    excel_path = "varga_result.xlsx"
    wb.save(excel_path)

    with open(excel_path, "rb") as f:
        st.download_button("📥 Excel 다운로드",
                           data=f,
                           file_name="varga_result.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


